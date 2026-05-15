from __future__ import annotations

from copy import copy
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


HEADER_CELL_MAP = {
    "sheet_title": "E2",
    "supervisor": "E3",
    "project_name": "E4",
    "collab_type": "E5",
    "brief_desc": "E6",
    "approval_date": "E7",
}

# Legacy mapping kept so older saved recipes can still be rendered.
LEGACY_CELL_MAP = {
    "substrate_overview": "D10",
    "substrate_details": "E10",
    "cleaning_overview": "D11",
    "cleaning_details": "E11",
    "resist1_overview": "D13",
    "resist1_details": "E13",
    "resist2_overview": "D14",
    "resist2_details": "E14",
    "exposure_overview": "D15",
    "exposure_details": "E15",
    "develop_overview": "D16",
    "develop_details": "E16",
    "deposition_overview": "D17",
    "deposition_details": "E17",
    "liftoff_overview": "D18",
    "liftoff_details": "E18",
    "inspection_overview": "D19",
    "inspection_details": "E19",
    "inspection_notes": "F19",
}

BLOCK_START_ROW = 10
BLOCK_TEMPLATE_END_ROW = 19
BLOCK_COLUMNS = {
    "numero": 2,
    "titulo": 3,
    "visao_geral": 4,
    "detalhes": 5,
    "notas": 6,
}


def _fmt_date(d: date | None):
    if d is None:
        return "---"
    return d.strftime("%d/%m/%Y")


def _copy_row_style(ws, source_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.number_format = source.number_format
        target.protection = copy(source.protection)


def _prepare_block_rows(ws, block_count: int):
    existing_rows = BLOCK_TEMPLATE_END_ROW - BLOCK_START_ROW + 1
    if block_count > existing_rows:
        ws.insert_rows(BLOCK_TEMPLATE_END_ROW + 1, block_count - existing_rows)

    total_rows = max(block_count, existing_rows)
    for idx in range(total_rows):
        row = BLOCK_START_ROW + idx
        source_row = min(row, BLOCK_TEMPLATE_END_ROW)
        if row > BLOCK_TEMPLATE_END_ROW:
            _copy_row_style(ws, source_row, row)
        for col in BLOCK_COLUMNS.values():
            ws.cell(row, col).value = None


def _write_blocks(ws, blocks: list[dict]):
    if not blocks:
        return

    _prepare_block_rows(ws, len(blocks))
    for idx, block in enumerate(blocks, start=1):
        row = BLOCK_START_ROW + idx - 1
        ws.cell(row, BLOCK_COLUMNS["numero"]).value = block.get("numero") or idx
        ws.cell(row, BLOCK_COLUMNS["titulo"]).value = block.get("titulo", "")
        ws.cell(row, BLOCK_COLUMNS["visao_geral"]).value = block.get("visao_geral", "")
        ws.cell(row, BLOCK_COLUMNS["detalhes"]).value = block.get("detalhes", "")
        ws.cell(row, BLOCK_COLUMNS["notas"]).value = block.get("notas", "")


def fill_sheet(template_path: Path, output_path: Path, payload: dict):
    wb = load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]

    for key, cell in HEADER_CELL_MAP.items():
        if key in payload:
            ws[cell] = payload.get(key, "")

    # If date was provided in a date widget.
    if isinstance(payload.get("approval_date"), date):
        ws[HEADER_CELL_MAP["approval_date"]] = _fmt_date(payload["approval_date"])

    blocks = payload.get("blocks") or []
    if blocks:
        _write_blocks(ws, blocks)
    else:
        for key, cell in LEGACY_CELL_MAP.items():
            if key in payload:
                ws[cell] = payload.get(key, "")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
